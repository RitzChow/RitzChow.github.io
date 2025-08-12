from __future__ import annotations

import json
import math
import os
from collections import defaultdict, OrderedDict
from dataclasses import dataclass
from typing import Dict, List, Any, Tuple

from flask import Flask, jsonify, send_from_directory
from openpyxl import load_workbook


# -------------------------
# Configuration
# -------------------------
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
INPUTS_CSV = os.path.join(PROJECT_ROOT, "ipho2025_answers.csv")
OUTPUTS_XLSX = os.path.join(PROJECT_ROOT, "ipho2025_outputs.xlsx")

COMPETITION_KEY = "ipho--ipho_2025"
COMPETITION_NICE_NAME = "IPhO 2025"


@dataclass
class RunRecord:
    problem_idx: str
    problem_statement: str
    model_name: str
    model_config: str
    idx_answer: int
    user_message: str
    answer: str
    messages: str
    input_tokens: float
    output_tokens: float
    run_cost: float
    input_cost_per_tokens: float
    output_cost_per_tokens: float
    gold_answer: str | None
    parsed_answer: str | None
    correct: bool | None


def read_inputs_csv(path: str) -> OrderedDict[str, str]:
    """Read mapping from sub-problem id -> gold answer in the provided order.
    Returns an OrderedDict preserving file order.
    """
    import csv
    ordered = OrderedDict()
    with open(path, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            pid = row.get("id")
            ans = row.get("answer")
            if pid:
                ordered[pid] = ans or ""
    return ordered


def read_outputs_xlsx(path: str) -> List[RunRecord]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    def get(row, key, default=None):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else default

    data: List[RunRecord] = []
    for row in rows[1:]:
        if not any(x is not None for x in row):
            continue
        try:
            record = RunRecord(
                problem_idx=str(get(row, "problem_idx")),
                problem_statement=str(get(row, "problem")),
                model_name=str(get(row, "model_name")),
                model_config=str(get(row, "model_config")),
                idx_answer=int(get(row, "idx_answer", 0) or 0),
                user_message=str(get(row, "user_message")),
                answer=str(get(row, "answer")),
                messages=str(get(row, "messages")),
                input_tokens=float(get(row, "input_tokens", 0) or 0),
                output_tokens=float(get(row, "output_tokens", 0) or 0),
                run_cost=float(get(row, "cost", 0) or 0),
                input_cost_per_tokens=float(get(row, "input_cost_per_tokens", 0) or 0),
                output_cost_per_tokens=float(get(row, "output_cost_per_tokens", 0) or 0),
                gold_answer=(get(row, "gold_answer") if get(row, "gold_answer") is not None else None),
                parsed_answer=(get(row, "parsed_answer") if get(row, "parsed_answer") is not None else None),
                correct=bool(get(row, "correct")) if get(row, "correct") is not None else None,
            )
            data.append(record)
        except Exception:
            # Skip malformed row but continue
            continue
    return data


def build_backend_payload() -> Tuple[Dict[str, Any], Dict[str, Any], Dict[str, Any], Dict[str, Any]]:
    # Inputs (gold answers) and ordering for problem names
    problem_id_to_gold = read_inputs_csv(INPUTS_CSV)
    problem_ids_in_order: List[str] = list(problem_id_to_gold.keys())

    runs = read_outputs_xlsx(OUTPUTS_XLSX)

    # Group runs: per (model_name, problem_id)
    grouped: Dict[Tuple[str, str], List[RunRecord]] = defaultdict(list)
    model_set: set[str] = set()
    for r in runs:
        model_set.add(r.model_name)
        grouped[(r.model_name, r.problem_idx)].append(r)

    # Aggregate per model totals for tokens and cost
    model_totals = {m: {"input_tokens": 0.0, "output_tokens": 0.0, "cost": 0.0} for m in model_set}
    # Capture per-token prices (divide by 1e6 if given in $/MTok)
    model_price: Dict[str, Dict[str, float]] = {m: {"input": None, "output": None} for m in model_set}

    for r in runs:
        mt = model_totals[r.model_name]
        mt["input_tokens"] += r.input_tokens or 0
        mt["output_tokens"] += r.output_tokens or 0
        mt["cost"] += r.run_cost or 0
        # Save first seen price. We display price per million tokens as-is from the spreadsheet.
        if model_price[r.model_name]["input"] is None and r.input_cost_per_tokens is not None:
            v = float(r.input_cost_per_tokens or 0)
            model_price[r.model_name]["input"] = v
        if model_price[r.model_name]["output"] is None and r.output_cost_per_tokens is not None:
            v = float(r.output_cost_per_tokens or 0)
            model_price[r.model_name]["output"] = v

    # Build results rows: one row per numeric task index, plus Avg and Cost rows
    # Map numeric column index -> problem id
    numeric_to_pid: Dict[int, str] = {}
    problem_names: List[str] = []
    for i, pid in enumerate(problem_ids_in_order, start=1):
        numeric_to_pid[i] = pid
        problem_names.append(pid)

    # For each numeric task, compute per-model accuracy %
    results_rows: List[Dict[str, Any]] = []
    for idx in range(1, len(problem_ids_in_order) + 1):
        pid = numeric_to_pid[idx]
        row: Dict[str, Any] = {"question": idx}
        for m in sorted(model_set):
            runs_for = grouped.get((m, pid), [])
            if not runs_for:
                row[m] = 0
            else:
                num = sum(1 for r in runs_for if (r.correct is True))
                den = len(runs_for)
                acc = 100.0 * num / den if den > 0 else 0.0
                row[m] = acc
        results_rows.append(row)

    # Avg row
    avg_row: Dict[str, Any] = {"question": "Avg"}
    for m in sorted(model_set):
        vals = [r[m] for r in results_rows if isinstance(r[m], (int, float))]
        avg_row[m] = sum(vals) / len(vals) if vals else 0.0
    results_rows.append(avg_row)

    # Cost row
    cost_row: Dict[str, Any] = {"question": "Cost"}
    for m in sorted(model_set):
        cost_row[m] = model_totals[m]["cost"]
    results_rows.append(cost_row)

    # Build top-level /results payload
    competition_info = {
        COMPETITION_KEY: {
            "index": 1,
            "nice_name": COMPETITION_NICE_NAME,
            "type": "FinalAnswer",
            "num_problems": len(problem_ids_in_order),
            "medal_thresholds": [75, 50, 25],
            "judge": False,
            "problem_names": problem_names,
        }
    }

    results_payload = {
        "competition_info": competition_info,
        "results": {
            COMPETITION_KEY: results_rows
        }
    }

    # Build /secondary payload rows
    secondary_rows: List[Dict[str, Any]] = []
    # Input Tokens
    row_it = {"question": "Input Tokens"}
    for m in sorted(model_set):
        row_it[m] = model_totals[m]["input_tokens"]
    secondary_rows.append(row_it)

    # Input Cost (total $): tokens * ($/Mtok) / 1e6
    row_icpt = {"question": "Input Cost"}
    for m in sorted(model_set):
        price_per_mtok = model_price[m]["input"] or 0.0
        tokens = model_totals[m]["input_tokens"] or 0.0
        dollars = (tokens * price_per_mtok) / 1_000_000.0
        row_icpt[m] = round(dollars, 6)
    secondary_rows.append(row_icpt)

    # Output Tokens
    row_ot = {"question": "Output Tokens"}
    for m in sorted(model_set):
        row_ot[m] = model_totals[m]["output_tokens"]
    secondary_rows.append(row_ot)

    # Output Cost (total $): tokens * ($/Mtok) / 1e6
    row_ocpt = {"question": "Output Cost"}
    for m in sorted(model_set):
        price_per_mtok = model_price[m]["output"] or 0.0
        tokens = model_totals[m]["output_tokens"] or 0.0
        dollars = (tokens * price_per_mtok) / 1_000_000.0
        row_ocpt[m] = round(dollars, 6)
    secondary_rows.append(row_ocpt)

    # Acc (overall average %)
    row_acc = {"question": "Acc"}
    for m in sorted(model_set):
        row_acc[m] = avg_row[m]
    secondary_rows.append(row_acc)

    secondary_payload = {COMPETITION_KEY: secondary_rows}

    # Build competition_dates: set all False (no contamination warning)
    competition_dates: Dict[str, Dict[str, bool]] = {COMPETITION_KEY: {}}
    for m in model_set:
        competition_dates[COMPETITION_KEY][m] = False

    # Build traces index: per (model, numeric_idx) produce trace structure
    traces_index: Dict[Tuple[str, int], Dict[str, Any]] = {}
    # For statement, take the first run's problem_statement per (model, pid)
    for idx in range(1, len(problem_ids_in_order) + 1):
        pid = numeric_to_pid[idx]
        gold = problem_id_to_gold.get(pid, "")
        for m in model_set:
            runs_for = sorted(grouped.get((m, pid), []), key=lambda r: r.idx_answer)
            if not runs_for:
                continue
            statement = runs_for[0].problem_statement or ""
            outs: List[Dict[str, Any]] = []
            for rr in runs_for:
                outs.append({
                    "parsed_answer": rr.parsed_answer,
                    "correct": bool(rr.correct) if rr.correct is not None else False,
                    "solution": rr.answer or "",
                })
            traces_index[(m, idx)] = {
                "statement": statement,
                "gold_answer": gold,
                "model_outputs": outs,
            }

    return results_payload, secondary_payload, competition_dates, traces_index


# -------------------------
# Flask app
# -------------------------
app = Flask(__name__, static_folder=PROJECT_ROOT, static_url_path="")

RESULTS_PAYLOAD, SECONDARY_PAYLOAD, COMP_DATES_PAYLOAD, TRACES_INDEX = build_backend_payload()


@app.route("/")
def root():
    return send_from_directory(PROJECT_ROOT, "index.html")


@app.get("/results")
def get_results():
    return jsonify(RESULTS_PAYLOAD)


@app.get("/secondary")
def get_secondary():
    return jsonify(SECONDARY_PAYLOAD)


@app.get("/competition_dates")
def get_comp_dates():
    return jsonify(COMP_DATES_PAYLOAD)


@app.get("/traces/<competition>/<model>/<int:task>")
def get_traces(competition: str, model: str, task: int):
    if competition != COMPETITION_KEY:
        return jsonify({"error": "competition not found"}), 404
    key = (model, task)
    data = TRACES_INDEX.get(key)
    if not data:
        return jsonify({"error": "trace not found"}), 404
    return jsonify(data)


def main():
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port, debug=True)


if __name__ == "__main__":
    main()


